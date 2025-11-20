from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.conf import settings
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.mail import get_connection, EmailMultiAlternatives
from django.utils import timezone
from django.contrib import messages
from django.contrib.staticfiles.storage import staticfiles_storage
from pathlib import Path
from datetime import datetime
from threading import Thread
import os, re

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .forms import ContactForm
from .utils_excel import append_submission_xlsx
from .utils_contact import normalize_phone_and_country, country_name_from_alpha2


# ---------- Validation patterns ----------
NAME_RE  = re.compile(r"^[A-Za-z\s'.-]{2,}$")
PHONE_RE = re.compile(r"^\+?\d[\d\s\-()]{6,}$")

# ---------- Excel paths ----------
EXCEL_DIR  = os.path.join(settings.BASE_DIR, "data")
EXCEL_PATH = os.path.join(EXCEL_DIR, "carl_demo_requests.xlsx")


# def _append_to_excel(row):
#     """Create/append to the Request Demo workbook."""
#     os.makedirs(EXCEL_DIR, exist_ok=True)
#     if os.path.exists(EXCEL_PATH):
#         wb = load_workbook(EXCEL_PATH)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Requests"
#         headers = [
#             "Timestamp", "Full Name", "Company", "Email",
#             "Country", "Dial Code", "Phone", "Address",
#             "Message", "Source IP",
#         ]
#         ws.append(headers)
#         for i in range(1, len(headers) + 1):
#             ws.column_dimensions[get_column_letter(i)].width = 24
#     ws.append(row)
#     wb.save(EXCEL_PATH)


# ---------- Email helpers ----------
def _send_email(subject: str, text_body: str, html_body: str | None, recipients: list[str] | None):
    """Low-level sender used by async wrappers."""
    try:
        if not recipients:
            # last-resort fallback
            fallback = getattr(settings, "EMAIL_HOST_USER", None) or getattr(settings, "DEFAULT_FROM_EMAIL", None)
            recipients = [fallback] if fallback else []

        if not recipients:
            print("EMAIL WARNING: no recipients configured")
            return

        conn = get_connection(timeout=getattr(settings, "EMAIL_TIMEOUT", 15))
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None) or getattr(settings, "EMAIL_HOST_USER", None),
            to=recipients,
            connection=conn,
        )
        if html_body:
            msg.attach_alternative(html_body, "text/html")
        msg.send(fail_silently=False)
    except Exception as e:
        print("EMAIL ERROR:", repr(e))


def _send_demo_email_async(subject: str, text_body: str, html_body: str | None = None):
    """Fire-and-forget email for Request Demo."""
    recipients = getattr(settings, "DEMO_RECIPIENTS", None) or getattr(settings, "CONTACT_RECIPIENTS", None)
    Thread(target=_send_email, args=(subject, text_body, html_body, recipients), daemon=True).start()


def _send_contact_email_async(subject: str, text_body: str, html_body: str | None = None):
    """Fire-and-forget email for Contact form."""
    recipients = getattr(settings, "CONTACT_RECIPIENTS", None)
    Thread(target=_send_email, args=(subject, text_body, html_body, recipients), daemon=True).start()


# ---------- Views ----------
def request_demo_view(request):
    if request.method != "POST":
        return redirect("/")

    # Pull fields
    full_name = request.POST.get("full_name", "").strip()
    company   = request.POST.get("company", "").strip()
    email     = request.POST.get("email", "").strip()
    phone     = request.POST.get("phone", "").strip()
    country   = request.POST.get("country", "").strip()  # "IN|+91"
    address   = request.POST.get("address", "").strip()
    message   = request.POST.get("message", "").strip()

    # Validate
    errors = {}
    if not NAME_RE.match(full_name):
        errors["full_name"] = "Please enter a valid full name (letters only)."
    if not company:
        errors["company"] = "Company is required."
    try:
        validate_email(email)
    except ValidationError:
        errors["email"] = "Enter a valid email."
    if not PHONE_RE.match(phone):
        errors["phone"] = "Enter a valid phone number."
    if not country:
        errors["country"] = "Select a country."

    if errors:
        # Raise toasts on next page load
        for msg in errors.values():
            messages.error(request, msg)
        # Go back to the page that opened the modal (so your JS toast can show)
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Split "IN|+91"
    country_code, dial = (country.split("|", 1) + [""])[:2]

    # Excel append
    # _append_to_excel([
    #     timezone.now().strftime("%Y-%m-%d %H:%M:%S %Z") or timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
    #     full_name, company, email, country_code, dial, phone, address, message,
    #     request.META.get("REMOTE_ADDR", ""),
    # ])

    # Build email
    ts = timezone.now().strftime("%Y-%m-%d %H:%M:%S %Z")
    subject = "New CARL Demo Request"
    text_body = (
        "A new CARL demo request was submitted.\n\n"
        f"Submitted: {ts}\n"
        f"IP: {request.META.get('REMOTE_ADDR','')}\n\n"
        f"Full name: {full_name}\n"
        f"Company: {company}\n"
        f"Email: {email}\n"
        f"Phone: {phone}\n"
        f"Country: {country_code} {dial}\n"
        f"Address: {address}\n\n"
        "Message:\n"
        f"{message or '(none)'}\n"
    )
    html_body = f"""
        <h2 style="margin:0 0 8px">New CARL Demo Request</h2>
        <p style="margin:0 0 12px;color:#334">Submitted {ts} from {request.META.get('REMOTE_ADDR','')}</p>
        <table cellpadding="6" cellspacing="0" style="border-collapse:collapse;background:#f9fbfc">
          <tr><td><b>Full name</b></td><td>{full_name}</td></tr>
          <tr><td><b>Company</b></td><td>{company}</td></tr>
          <tr><td><b>Email</b></td><td>{email}</td></tr>
          <tr><td><b>Phone</b></td><td>{phone}</td></tr>
          <tr><td><b>Country</b></td><td>{country_code} {dial}</td></tr>
          <tr><td><b>Address</b></td><td>{address}</td></tr>
        </table>
        <p style="margin:12px 0 4px"><b>Message</b></p>
        <pre style="white-space:pre-wrap;font-family:system-ui,Segoe UI,Arial,sans-serif">{message or '(none)'}</pre>
    """

    _send_demo_email_async(subject, text_body, html_body)

    # Success -> thanks page
    return redirect(reverse("cmmsApp:contact_thanks"))


def home(request):
    return render(request, "index.html")


def sitemap(request):
    with staticfiles_storage.open('sitemap.xml') as sitemap_file:
        return HttpResponse(sitemap_file, content_type='application/xml')
    
    
def request_demo(request):
    return render(request, "request_demo_modal.html")
def factory(request):     return render(request, "factory.html")

def factory(request):     return render(request, "factory.html")
def healthcare(request):  return render(request, "healthcare.html")
def facility(request):    return render(request, "facility.html")
def city(request):        return render(request, "city.html")
def transport(request):   return render(request, "transport.html")
def contact(request):     return render(request, "contact.html")
def iot(request):         return render(request, "iot.html")
def eam(request):         return render(request, "eam.html")
def apm(request):         return render(request, "apm.html")
def mobility(request):    return render(request, "mobility.html")
def plans(request):       return render(request, "plans.html")
def about(request):       return render(request, "about.html")
def workorder(request):   return render(request, "workorder.html")
def compliance(request):  return render(request, "compliance.html")
def cmmsiot(request):       return render(request, "cmms-iot.html")
def gis(request):         return render(request, "gis.html")
def erpsync(request):     return render(request, "erpsync.html")
def industries(request):     return render(request, "industries.html")


def contact_section(request):
    form = ContactForm(request.POST or None)

    if request.method == "POST" and not form.is_valid():
        messages.error(request, "Please correct the highlighted fields and resubmit.")

    if request.method == "POST" and form.is_valid():
        cd = form.cleaned_data

        # Normalize phone & resolve country name
        e164_phone, resolved_alpha2, resolved_country_name = normalize_phone_and_country(
            cd.get("phone", ""), cd.get("country", "")
        )

        # Append to Excel
        # xlsx_path = Path(
        #     getattr(settings, "CONTACT_SUBMISSIONS_XLSX", Path(settings.BASE_DIR) / "contact_submissions.xlsx")
        # )
        # append_submission_xlsx(
        #     xlsx_path,
        #     [
        #         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        #         cd["first_name"],
        #         cd.get("last_name", ""),
        #         cd.get("company", ""),
        #         cd["email"],
        #         resolved_alpha2,
        #         resolved_country_name,
        #         e164_phone or cd.get("phone", ""),
        #         cd.get("message", ""),
        #     ],
        # )

        # Email body
        subject = "New website contact submission for CARL Software"
        text_body = "\n".join(
            [
                "New contact submission for CARL Software:",
                f"Name: {cd['first_name']} {cd.get('last_name','')}".strip(),
                f"Company: {cd.get('company','')}",
                f"Email: {cd['email']}",
                f"Country: {resolved_country_name or country_name_from_alpha2(resolved_alpha2) or cd.get('country','')}",
                f"Phone: {e164_phone or cd.get('phone','')}",
                "",
                "Message:",
                cd.get("message", ""),
            ]
        )

        _send_contact_email_async(subject, text_body, None)

        return redirect(reverse("cmmsApp:contact_thanks"))

    return render(request, "contact_section.html", {"form": form, "sent": request.GET.get("sent")})


def contact_thanks(request):
    return render(request, "contact_thanks.html", {})
